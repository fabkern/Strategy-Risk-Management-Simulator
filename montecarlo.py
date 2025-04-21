import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill
from openpyxl.drawing.image import Image
import openpyxl

# Input parameters
initial_capital = 1000  # Initial capital ($)
win_rate = 0.10  # 0.55% = 55% Win rate as a percentage
risk_to_reward = 15  # Risk to reward ratio (1:X)
percent_balance = None  # 1% = 1% Percentage of balance (None if using fixed dollar)
fixed_amount = 10  # Fixed dollar amount ($)
num_trades = 800  # Number of trades
concurrent_trades = 20  # Number of concurrent trades

# Function to run Monte Carlo simulation
def monte_carlo_simulation(win_rate, risk_to_reward, percent_balance, fixed_amount, num_trades, concurrent_trades, initial_capital):
    balances = []
    trade_results = []
    balance = initial_capital

    for _ in range(num_trades):
        rand = np.random.rand()  # Random number to simulate win or loss

        # Adjust risk for concurrent trades if percentage of balance is used
        adjusted_percent_balance = percent_balance / concurrent_trades if percent_balance else None

        if rand <= win_rate:  # Win
            if percent_balance:
                profit = (adjusted_percent_balance / 100) * balance * risk_to_reward
            else:
                profit = fixed_amount * risk_to_reward
            balance += profit
            trade_results.append(1)  # Win
        else:  # Loss
            if percent_balance:
                loss = -(adjusted_percent_balance / 100) * balance
            else:
                loss = -fixed_amount
            balance += loss
            trade_results.append(0)  # Loss

        balances.append(balance)

    return np.array(balances), trade_results


# Function to calculate max consecutive wins and losses
def calculate_consecutive_wins_losses(trade_results):
    max_consecutive_wins = 0
    max_consecutive_losses = 0
    current_wins = 0
    current_losses = 0

    for result in trade_results:
        if result == 1:  # Win
            current_wins += 1
            current_losses = 0
        else:  # Loss
            current_losses += 1
            current_wins = 0

        max_consecutive_wins = max(max_consecutive_wins, current_wins)
        max_consecutive_losses = max(max_consecutive_losses, current_losses)

    return max_consecutive_wins, max_consecutive_losses

# Function to run the simulation and generate reports
def generate_report(initial_capital, win_rate, risk_to_reward, percent_balance=None, fixed_amount=None, num_trades=100, concurrent_trades=1):
    best_case = None
    worst_case = None
    most_likely_case = None
    all_scenarios = []
    best_results = []
    worst_results = []
    most_likely_results = []

    for i in range(1000):  # Running 1000 simulations
        balances, results = monte_carlo_simulation(win_rate, risk_to_reward, percent_balance, fixed_amount, num_trades, concurrent_trades, initial_capital)
        all_scenarios.append(balances)

        if best_case is None or balances[-1] > best_case[-1]:
            best_case = balances
            best_results = results

        if worst_case is None or balances[-1] < worst_case[-1]:
            worst_case = balances
            worst_results = results

        # Calculate most likely as median
        all_scenarios_np = np.array(all_scenarios)
        most_likely_case = np.median(all_scenarios_np, axis=0)

    most_likely_results = monte_carlo_simulation(win_rate, risk_to_reward, percent_balance, fixed_amount, num_trades, concurrent_trades, initial_capital)[1]

    return best_case, worst_case, most_likely_case, all_scenarios, best_results, worst_results, most_likely_results

# Function to write results to Excel
def write_to_excel(best, worst, most_likely, all_scenarios, initial_capital, num_trades, best_results, worst_results, most_likely_results):
    wb = Workbook()

    # First sheet: All simulations
    ws1 = wb.active
    ws1.title = "All Simulations"
    df = pd.DataFrame(all_scenarios)
    for r in dataframe_to_rows(df, index=False, header=False):
        ws1.append(r)

    # Second sheet: Best, Worst, and Most Likely scenarios with plot
    ws2 = wb.create_sheet("Best, Worst, Most Likely")

    df_plot = pd.DataFrame({"Best": best, "Worst": worst, "Most Likely": most_likely})

    # Add the data
    for r in dataframe_to_rows(df_plot, index=False, header=True):
        ws2.append(r)

    # Plotting and saving chart inside the Excel
    plt.figure(figsize=(10, 6))
    plt.plot(best, label='Best (Green)', color='green')
    plt.plot(worst, label='Worst (Red)', color='red')
    plt.plot(most_likely, label='Most Likely (Blue)', color='blue')
    plt.legend()
    plt.title("Best, Worst, and Most Likely Scenarios")
    plt.xlabel("Trade Number")
    plt.ylabel("Balance ($)")
    plt.grid(True)
    plt.tight_layout()
    plt.savefig("monte_carlo_plot.png")

    img = openpyxl.drawing.image.Image("monte_carlo_plot.png")
    ws2.add_image(img, 'F2')

    # Third sheet: Summary of Results
    ws3 = wb.create_sheet("Summary")

    max_consecutive_wins_best, max_consecutive_losses_best = calculate_consecutive_wins_losses(best_results)
    max_consecutive_wins_worst, max_consecutive_losses_worst = calculate_consecutive_wins_losses(worst_results)
    max_consecutive_wins_most, max_consecutive_losses_most = calculate_consecutive_wins_losses(most_likely_results)

    summary_data = {
        'Scenario': ['Best', 'Worst', 'Most Likely'],
        'Start Balance ($)': [initial_capital, initial_capital, initial_capital],
        'End Balance ($)': [best[-1], worst[-1], most_likely[-1]],
        'Results in %': [(best[-1] - initial_capital) / initial_capital * 100,
                         (worst[-1] - initial_capital) / initial_capital * 100,
                         (most_likely[-1] - initial_capital) / initial_capital * 100],
        'Max Drawdown (%)': [
            (initial_capital - min(best)) / initial_capital * 100,
            (initial_capital - min(worst)) / initial_capital * 100,
            (initial_capital - min(most_likely)) / initial_capital * 100
        ],
        'Max Drawdown ($)': [
            initial_capital - min(best),
            initial_capital - min(worst),
            initial_capital - min(most_likely)
        ],
        'Max Consecutive Losses': [max_consecutive_losses_best, max_consecutive_losses_worst, max_consecutive_losses_most],
        'Max Consecutive Wins': [max_consecutive_wins_best, max_consecutive_wins_worst, max_consecutive_wins_most]
    }

    df_summary = pd.DataFrame(summary_data)

    for r in dataframe_to_rows(df_summary, index=False, header=True):
        ws3.append(r)

    # Coloring best, worst, and most likely cells
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="none")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="none")
    blue_fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="none")

    for cell in ws3['A']:
        if cell.value == 'Best':
            for col in range(2, 8):
                ws3.cell(row=cell.row, column=col).fill = green_fill
        elif cell.value == 'Worst':
            for col in range(2, 8):
                ws3.cell(row=cell.row, column=col).fill = red_fill
        elif cell.value == 'Most Likely':
            for col in range(2, 8):
                ws3.cell(row=cell.row, column=col).fill = blue_fill

    # Save the workbook
    wb.save(f"Monte_Carlo_Simulation_WR_{win_rate}_RR_{risk_to_reward}.xlsx")


# Run the simulation and generate the report
best, worst, most_likely, all_scenarios, best_results, worst_results, most_likely_results = generate_report(
    initial_capital, win_rate, risk_to_reward, percent_balance, fixed_amount, num_trades, concurrent_trades)

write_to_excel(best, worst, most_likely, all_scenarios, initial_capital, num_trades, best_results, worst_results, most_likely_results)
